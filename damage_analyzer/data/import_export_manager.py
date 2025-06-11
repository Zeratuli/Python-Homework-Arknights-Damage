# import_export_manager.py - 统一的导入导出管理器

import os
import json
import csv
import logging
import time
from typing import Dict, List, Any, Optional, Tuple, Callable
from tkinter import messagebox, filedialog
from datetime import datetime

# 导入现有的处理器
from .excel_handler import ExcelHandler
from .json_handler import JsonHandler
from .csv_handler import CsvHandler

logger = logging.getLogger(__name__)

class ImportExportManager:
    """统一的导入导出管理器"""
    
    def __init__(self, db_manager):
        """
        初始化导入导出管理器
        
        Args:
            db_manager: 数据库管理器实例
        """
        self.db_manager = db_manager
        self.status_callback = None
        self.refresh_callback = None  # 新增：刷新通知回调
        self.overview_panel = None    # 新增：概览面板引用
        
        # 创建日志记录器
        self.logger = logging.getLogger(__name__)
        
        # 支持的文件类型
        self.supported_formats = {
            'excel': ['.xlsx', '.xls'],
            'json': ['.json'],
            'csv': ['.csv']
        }
        
        # 初始化处理器
        self.excel_handler = ExcelHandler()
        self.json_handler = JsonHandler()
        self.csv_handler = CsvHandler()
    
    def set_status_callback(self, callback: Callable[[str, str], None]):
        """设置状态回调函数"""
        self.status_callback = callback
    
    def set_refresh_callback(self, callback: Callable[[], None]):
        """设置刷新通知回调函数"""
        self.refresh_callback = callback
    
    def set_overview_panel(self, overview_panel):
        """设置概览面板引用，用于实时活动推送"""
        self.overview_panel = overview_panel
    
    def _update_status(self, message: str, level: str = "info"):
        """更新状态"""
        if self.status_callback:
            self.status_callback(message, level)
    
    def _notify_refresh(self, import_type: str, import_count: int):
        """通知界面刷新统计数据"""
        try:
            self.logger.info(f"开始通知刷新: {import_type}, 导入数量: {import_count}")
            
            # 确保数据库事务已完成，稍微延迟刷新
            time.sleep(0.1)  # 延迟100ms确保数据库事务完成
            
            if self.refresh_callback:
                self.logger.info("调用刷新回调...")
                self.refresh_callback()
                self.logger.info("刷新回调调用完成")
            else:
                self.logger.warning("刷新回调未设置")
            
            # 推送实时活动记录
            self._push_activity_record(f"导入了{import_count}条{import_type}数据")
            
            self.logger.info(f"刷新通知完成: {import_type}")
            
        except Exception as e:
            # 刷新失败不影响主要功能
            self.logger.error(f"通知刷新失败: {e}")
    
    def _push_activity_record(self, activity_description: str):
        """推送实时活动记录"""
        try:
            self.logger.info(f"推送活动记录: {activity_description}")
            
            # 如果有概览面板引用，直接推送
            if self.overview_panel and hasattr(self.overview_panel, 'push_real_time_activity'):
                try:
                    self.overview_panel.push_real_time_activity(activity_description)
                    self.logger.info(f"直接推送活动记录成功: {activity_description}")
                    return
                except Exception as e:
                    self.logger.error(f"直接推送活动记录失败: {e}")
            
            # 降级方案：使用临时文件机制
            activity_file = "temp_activity.txt"
            try:
                with open(activity_file, "w", encoding="utf-8") as f:
                    f.write(f"{activity_description}")
                self.logger.info(f"活动记录已写入临时文件: {activity_file}")
            except Exception as file_error:
                self.logger.error(f"写入活动记录文件失败: {file_error}")
                
        except Exception as e:
            self.logger.error(f"推送活动记录失败: {e}")
            pass
    
    def import_excel_data(self, filename: str = None, status_callback: Callable = None) -> Dict[str, Any]:
        """
        导入Excel数据
        
        Args:
            filename: Excel文件路径
            status_callback: 状态回调函数
            
        Returns:
            导入结果字典
        """
        try:
            if status_callback:
                self.set_status_callback(status_callback)
            
            # 使用Excel处理器导入数据
            operators, errors = self.excel_handler.import_from_excel(filename)
            
            if not operators and not errors:
                # 用户取消了文件选择
                return {'success': False, 'cancelled': True}
            
            if errors:
                # 显示错误信息
                error_msg = "\n".join(errors[:10])  # 最多显示10个错误
                if len(errors) > 10:
                    error_msg += f"\n... 还有 {len(errors) - 10} 个错误"
                
                if not operators:
                    messagebox.showerror("导入失败", f"Excel导入失败：\n{error_msg}")
                    return {'success': False, 'error': error_msg}
                else:
                    # 有部分成功
                    result = messagebox.askyesno(
                        "部分导入成功", 
                        f"导入过程中发现以下错误：\n{error_msg}\n\n"
                        f"成功解析 {len(operators)} 个干员数据，是否继续导入这些数据？"
                    )
                    if not result:
                        return {'success': False, 'cancelled': True}
            
            # 导入到数据库
            import_count = 0
            duplicate_count = 0
            
            for operator in operators:
                try:
                    # 检查是否已存在
                    existing = self.db_manager.get_operator_by_name(operator['name'])
                    if existing:
                        # 询问是否覆盖
                        result = messagebox.askyesno(
                            "重复数据", 
                            f"干员 '{operator['name']}' 已存在，是否覆盖？"
                        )
                        if result:
                            try:
                                # 尝试更新干员，捕获具体错误
                                success = self.db_manager.update_operator(existing['id'], operator)
                                if success:
                                    import_count += 1
                                    self.logger.info(f"成功覆盖干员: {operator['name']} (ID: {existing['id']})")
                                else:
                                    self.logger.warning(f"覆盖干员失败，没有找到记录: {operator['name']}")
                                    duplicate_count += 1
                            except ValueError as ve:
                                # 名称冲突等逻辑错误
                                self.logger.error(f"覆盖干员失败 - 约束冲突: {ve}")
                                messagebox.showerror(
                                    "覆盖失败", 
                                    f"无法覆盖干员 '{operator['name']}'：\n{str(ve)}"
                                )
                                duplicate_count += 1
                            except Exception as update_error:
                                # 其他数据库错误
                                self.logger.error(f"覆盖干员失败 - 数据库错误: {update_error}")
                                messagebox.showerror(
                                    "数据库错误", 
                                    f"覆盖干员 '{operator['name']}' 时发生错误：\n{str(update_error)}"
                                )
                                duplicate_count += 1
                        else:
                            duplicate_count += 1
                    else:
                        try:
                            # 尝试插入新干员
                            operator_id = self.db_manager.insert_operator(operator)
                            import_count += 1
                            self.logger.info(f"成功插入新干员: {operator['name']} (ID: {operator_id})")
                        except Exception as insert_error:
                            self.logger.error(f"插入干员失败: {insert_error}")
                            # 显示更具体的错误信息
                            messagebox.showerror(
                                "插入失败", 
                                f"插入干员 '{operator['name']}' 时发生错误：\n{str(insert_error)}"
                            )
                        
                except Exception as e:
                    self.logger.warning(f"导入干员 {operator.get('name', '未知')} 失败: {e}")
                    # 显示更具体的错误信息
                    messagebox.showwarning(
                        "导入警告", 
                        f"处理干员 '{operator.get('name', '未知')}' 时出现问题：\n{str(e)}"
                    )
            
            # 记录导入操作 - 修复：无论filename是否为None都记录导入操作
            try:
                file_name = os.path.basename(filename) if filename else "手动导入"
                self.db_manager.record_import(
                    import_type='excel_import',
                    file_name=file_name,
                    record_count=import_count,
                    status='success' if import_count > 0 else 'partial'
                )
                self.logger.info(f"已记录导入操作: 文件={file_name}, 数量={import_count}")
            except Exception as record_error:
                self.logger.error(f"记录导入操作失败: {record_error}")
                # 即使记录失败，也不影响导入流程
            
            # 构建结果消息
            result_msg = f"Excel导入完成！\n"
            result_msg += f"成功导入: {import_count} 个干员\n"
            if duplicate_count > 0:
                result_msg += f"跳过重复: {duplicate_count} 个干员\n"
            if errors:
                result_msg += f"解析错误: {len(errors)} 个"
            
            self._update_status(f"Excel导入成功: {import_count} 个干员", "success")
            
            self._notify_refresh('excel', import_count)
            
            return {
                'success': True,
                'import_count': import_count,
                'duplicate_count': duplicate_count,
                'error_count': len(errors),
                'message': result_msg
            }
            
        except ImportError:
            error_msg = "需要安装pandas和openpyxl库才能导入Excel文件\n请运行: pip install pandas openpyxl"
            messagebox.showerror("错误", error_msg)
            return {'success': False, 'error': error_msg}
        except Exception as e:
            logger.error(f"Excel导入失败: {e}")
            error_msg = f"Excel导入失败：\n{str(e)}"
            messagebox.showerror("导入失败", error_msg)
            self._update_status("Excel导入失败", "error")
            return {'success': False, 'error': str(e)}
    
    def import_json_data(self, filename: str = None, status_callback: Callable = None) -> Dict[str, Any]:
        """
        导入JSON数据
        
        Args:
            filename: JSON文件路径
            status_callback: 状态回调函数
            
        Returns:
            导入结果字典
        """
        try:
            if status_callback:
                self.set_status_callback(status_callback)
            
            if not filename:
                filename = filedialog.askopenfilename(
                    title="导入JSON文件",
                    filetypes=[("JSON 文件", "*.json")]
                )
                
            if not filename:
                return {'success': False, 'cancelled': True}
            
            import_count = 0
            duplicate_count = 0
            errors = []
            
            with open(filename, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # 处理不同的JSON格式
            if isinstance(data, list):
                operators_data = data
            elif isinstance(data, dict):
                # 尝试找到干员数据数组
                operators_data = data.get('operators', data.get('data', [data]))
            else:
                raise ValueError("不支持的JSON格式")
            
            # 字段映射表（支持多种命名方式）
            field_mappings = {
                'name': ['name', '名称', '干员名称', '姓名'],
                'class_type': ['class_type', '职业', '职业类型', 'class', 'type'],
                'atk': ['atk', 'attack', '攻击力', '攻击'],
                'hp': ['hp', 'health', '生命值', '血量', '生命'],
                'def': ['def', 'defense', '防御', '防御力'],
                'atk_speed': ['atk_speed', 'attack_speed', '攻击速度', '攻速'],
                'cost': ['cost', '费用', '部署费用', 'deploy_cost'],
                'block_count': ['block_count', 'block', '阻挡', '阻挡数'],
                'atk_type': ['atk_type', 'attack_type', '攻击类型', '伤害类型']
            }
            
            for item in operators_data:
                try:
                    operator = {}
                    
                    # 模糊字段映射
                    for db_field, possible_names in field_mappings.items():
                        for name in possible_names:
                            if name in item and item[name] is not None:
                                operator[db_field] = item[name]
                                break
                    
                    # 验证必要字段
                    if not operator.get('name'):
                        errors.append(f"缺少干员名称: {item}")
                        continue
                        
                    # 设置默认值
                    operator.setdefault('class_type', '未知')
                    operator.setdefault('atk', 0)
                    operator.setdefault('hp', 0)
                    operator.setdefault('def', 0)
                    operator.setdefault('atk_speed', 1.0)
                    operator.setdefault('cost', 10)
                    operator.setdefault('block_count', 1)
                    operator.setdefault('atk_type', '物伤')
                    
                    # 检查是否已存在
                    existing = self.db_manager.get_operator_by_name(operator['name'])
                    if existing:
                        result = messagebox.askyesno(
                            "重复数据", 
                            f"干员 '{operator['name']}' 已存在，是否覆盖？"
                        )
                        if result:
                            try:
                                # 尝试更新干员，捕获具体错误
                                success = self.db_manager.update_operator(existing['id'], operator)
                                if success:
                                    import_count += 1
                                    self.logger.info(f"成功覆盖干员: {operator['name']} (ID: {existing['id']})")
                                else:
                                    self.logger.warning(f"覆盖干员失败，没有找到记录: {operator['name']}")
                                    duplicate_count += 1
                            except ValueError as ve:
                                # 名称冲突等逻辑错误
                                self.logger.error(f"覆盖干员失败 - 约束冲突: {ve}")
                                messagebox.showerror(
                                    "覆盖失败", 
                                    f"无法覆盖干员 '{operator['name']}'：\n{str(ve)}"
                                )
                                duplicate_count += 1
                            except Exception as update_error:
                                # 其他数据库错误
                                self.logger.error(f"覆盖干员失败 - 数据库错误: {update_error}")
                                messagebox.showerror(
                                    "数据库错误", 
                                    f"覆盖干员 '{operator['name']}' 时发生错误：\n{str(update_error)}"
                                )
                                duplicate_count += 1
                        else:
                            duplicate_count += 1
                    else:
                        try:
                            # 尝试插入新干员
                            operator_id = self.db_manager.insert_operator(operator)
                            import_count += 1
                            self.logger.info(f"成功插入新干员: {operator['name']} (ID: {operator_id})")
                        except Exception as insert_error:
                            self.logger.error(f"插入干员失败: {insert_error}")
                            # 显示更具体的错误信息
                            messagebox.showerror(
                                "插入失败", 
                                f"插入干员 '{operator['name']}' 时发生错误：\n{str(insert_error)}"
                            )
                        
                except Exception as e:
                    self.logger.warning(f"导入干员 {operator.get('name', '未知')} 失败: {e}")
                    # 显示更具体的错误信息
                    messagebox.showwarning(
                        "导入警告", 
                        f"处理干员 '{operator.get('name', '未知')}' 时出现问题：\n{str(e)}"
                    )
            
            # 记录导入操作 - 修复：JSON导入也要确保记录
            try:
                file_name = os.path.basename(filename) if filename else "手动导入"
                self.db_manager.record_import(
                    import_type='json_import',
                    file_name=file_name,
                    record_count=import_count,
                    status='success' if import_count > 0 else 'partial'
                )
                self.logger.info(f"已记录JSON导入操作: 文件={file_name}, 数量={import_count}")
            except Exception as record_error:
                self.logger.error(f"记录JSON导入操作失败: {record_error}")
                # 即使记录失败，也不影响导入流程
            
            # 构建结果消息
            result_msg = f"JSON导入完成！\n"
            result_msg += f"成功导入: {import_count} 个干员\n"
            if duplicate_count > 0:
                result_msg += f"跳过重复: {duplicate_count} 个干员\n"
            if errors:
                result_msg += f"处理错误: {len(errors)} 个"
            
            self._update_status(f"JSON导入成功: {import_count} 个干员", "success")
            
            self._notify_refresh('json', import_count)
            
            return {
                'success': True,
                'import_count': import_count,
                'duplicate_count': duplicate_count,
                'error_count': len(errors),
                'message': result_msg
            }
            
        except Exception as e:
            logger.error(f"JSON导入失败: {e}")
            error_msg = f"JSON导入失败：\n{str(e)}"
            messagebox.showerror("导入失败", error_msg)
            self._update_status("JSON导入失败", "error")
            return {'success': False, 'error': str(e)}
    
    def import_csv_data(self, filename: str = None, status_callback: Callable = None) -> Dict[str, Any]:
        """
        导入CSV数据
        
        Args:
            filename: CSV文件路径
            status_callback: 状态回调函数
            
        Returns:
            导入结果字典
        """
        try:
            if status_callback:
                self.set_status_callback(status_callback)
            
            if not filename:
                filename = filedialog.askopenfilename(
                    title="导入CSV文件",
                    filetypes=[("CSV 文件", "*.csv")]
                )
                
            if not filename:
                return {'success': False, 'cancelled': True}
            
            import_count = 0
            update_count = 0
            
            with open(filename, 'r', encoding='utf-8-sig') as csvfile:  # 使用utf-8-sig自动处理BOM
                reader = csv.DictReader(csvfile)
                for row in reader:
                    # 根据实际CSV格式进行数据映射 - 修复：处理BOM和优先使用中文列名
                    operator_data = {}
                    
                    # 处理可能的BOM问题，清理所有key
                    cleaned_row = {}
                    for key, value in row.items():
                        # 移除BOM字符
                        clean_key = key.lstrip('\ufeff').strip()
                        cleaned_row[clean_key] = value.strip() if isinstance(value, str) else value
                    
                    # 名称字段 - 修复：优先检查中文列名
                    operator_data['name'] = cleaned_row.get('名称', cleaned_row.get('干员名称', cleaned_row.get('name', '')))
                    
                    # 职业类型 - 修复：优先检查中文列名
                    operator_data['class_type'] = cleaned_row.get('职业类型', cleaned_row.get('职业', cleaned_row.get('class_type', '')))
                    
                    # 数值字段 - 修复：优先检查中文列名
                    try:
                        operator_data['atk'] = int(float(cleaned_row.get('攻击力', cleaned_row.get('atk', 0))))
                        operator_data['hp'] = int(float(cleaned_row.get('生命值', cleaned_row.get('hp', 0))))
                        operator_data['def'] = int(float(cleaned_row.get('防御力', cleaned_row.get('def', 0))))
                        operator_data['mdef'] = int(float(cleaned_row.get('法抗', cleaned_row.get('mdef', 0))))
                        operator_data['cost'] = int(float(cleaned_row.get('部署费用', cleaned_row.get('费用', cleaned_row.get('cost', 10)))))
                        operator_data['block_count'] = int(float(cleaned_row.get('阻挡数', cleaned_row.get('阻挡', cleaned_row.get('block_count', 1)))))
                        operator_data['atk_speed'] = float(cleaned_row.get('攻击速度', cleaned_row.get('atk_speed', 1.0)))
                    except (ValueError, TypeError):
                        # 如果数值转换失败，使用默认值
                        operator_data.setdefault('atk', 0)
                        operator_data.setdefault('hp', 0)
                        operator_data.setdefault('def', 0)
                        operator_data.setdefault('mdef', 0)
                        operator_data.setdefault('cost', 10)
                        operator_data.setdefault('block_count', 1)
                        operator_data.setdefault('atk_speed', 1.0)
                    
                    # 攻击类型 - 修复：优先检查中文列名
                    operator_data['atk_type'] = cleaned_row.get('攻击类型', cleaned_row.get('伤害类型', cleaned_row.get('atk_type', '物理伤害')))
                    
                    if operator_data['name']:  # 只导入有名称的干员
                        # 检查是否已存在
                        existing = self.db_manager.get_operator_by_name(operator_data['name'])
                        if existing:
                            # 默认覆盖，不询问用户
                            try:
                                success = self.db_manager.update_operator(existing['id'], operator_data)
                                if success:
                                    update_count += 1
                                    self.logger.info(f"CSV导入：覆盖干员 {operator_data['name']} (ID: {existing['id']})")
                                else:
                                    self.logger.warning(f"CSV导入：覆盖干员失败 {operator_data['name']}")
                            except Exception as update_error:
                                self.logger.error(f"CSV导入：覆盖干员异常 {operator_data['name']}: {update_error}")
                        else:
                            # 插入新干员
                            try:
                                operator_id = self.db_manager.insert_operator(operator_data)
                                import_count += 1
                                self.logger.info(f"CSV导入：新增干员 {operator_data['name']} (ID: {operator_id})")
                            except Exception as insert_error:
                                self.logger.error(f"CSV导入：新增干员异常 {operator_data['name']}: {insert_error}")
            
            # 记录导入 - 修复：CSV导入也要确保记录
            try:
                file_name = os.path.basename(filename) if filename else "手动导入"
                total_processed = import_count + update_count
                self.db_manager.record_import(
                    import_type='csv_import',
                    file_name=file_name,
                    record_count=total_processed,
                    status='success'
                )
                self.logger.info(f"已记录CSV导入操作: 文件={file_name}, 新增={import_count}, 更新={update_count}")
            except Exception as record_error:
                self.logger.error(f"记录CSV导入操作失败: {record_error}")
                # 即使记录失败，也不影响导入流程
            
            # 构建详细的结果消息
            if update_count > 0:
                result_msg = f"CSV导入成功！新增 {import_count} 个，更新 {update_count} 个干员"
            else:
                result_msg = f"CSV导入成功！新增 {import_count} 个干员"
            
            total_processed = import_count + update_count
            self._update_status(f"CSV导入成功: 新增{import_count}个，更新{update_count}个", "success")
            
            self._notify_refresh('csv', total_processed)
            
            return {
                'success': True,
                'import_count': import_count,
                'update_count': update_count,
                'total_processed': total_processed,
                'message': result_msg
            }
            
        except Exception as e:
            logger.error(f"CSV导入失败: {e}")
            error_msg = f"CSV导入失败：\n{str(e)}"
            messagebox.showerror("导入失败", error_msg)
            self._update_status("CSV导入失败", "error")
            return {'success': False, 'error': str(e)}
    
    def export_current_data(self, operators: List[Dict[str, Any]], filename: str = None, current_charts: List[Dict] = None) -> bool:
        """
        导出当前数据（支持包含用户生成的图表）
        
        Args:
            operators: 干员数据列表
            filename: 导出文件路径
            current_charts: 用户生成的图表列表
            
        Returns:
            是否导出成功
        """
        try:
            if not operators:
                messagebox.showwarning("警告", "没有数据可导出")
                return False
            
            # 获取计算结果：优先获取用户当前的计算结果，然后补充数据库记录
            recent_calculations = self._get_current_and_recent_calculations()
            
            if filename is None:
                filename = filedialog.asksaveasfilename(
                    title="导出当前数据",
                    defaultextension=".json",
                    filetypes=[
                        ("JSON 文件", "*.json"),
                        ("CSV 文件", "*.csv"),
                        ("Excel 文件", "*.xlsx"),
                        ("HTML 报告", "*.html"),
                        ("所有文件", "*.*")
                    ]
                )
                
            if not filename:
                return False
            
            file_ext = filename.lower().split('.')[-1]
            base_filename = filename.rsplit('.', 1)[0]  # 不包含扩展名的文件名
            
            # 保存用户生成的图表为图片文件
            chart_paths = []
            if current_charts:
                chart_paths = self._save_current_charts_as_images(current_charts, base_filename)
                logger.info(f"保存了 {len(chart_paths)} 个用户生成的图表")
            else:
                logger.info("没有用户生成的图表可导出")
            
            if file_ext == 'json':
                # 导出为JSON格式（包含计算结果）
                export_data = {
                    'operators': operators,
                    'recent_calculations': recent_calculations,
                    'export_info': {
                        'export_time': datetime.now().isoformat(),
                        'total_operators': len(operators),
                        'chart_count': len(chart_paths),
                        'calculation_count': len(recent_calculations),
                        'has_user_charts': len(chart_paths) > 0
                    }
                }
                
                with open(filename, 'w', encoding='utf-8') as f:
                    json.dump(export_data, f, ensure_ascii=False, indent=2)
                    
            elif file_ext == 'csv':
                # 导出为CSV格式 + 图表 + 计算结果
                self._export_csv_with_charts_and_calculations(operators, filename, chart_paths, recent_calculations)
                
            elif file_ext in ['xlsx', 'xls']:
                # 导出为Excel格式 + 图表 + 计算结果
                self._export_excel_with_charts_and_calculations(operators, filename, chart_paths, recent_calculations)
                
            elif file_ext == 'html':
                # 导出为HTML报告 + 内嵌图表 + 计算结果
                self._export_html_with_charts_and_calculations(operators, filename, chart_paths, recent_calculations)
            
            # 记录导出操作
            self.db_manager.record_import(
                import_type='export_with_user_charts_and_calculations',
                file_name=os.path.basename(filename),
                record_count=len(operators),
                status='success'
            )
            
            # 显示成功消息
            chart_info = f"\n包含 {len(chart_paths)} 个用户生成的图表" if chart_paths else "\n未包含图表（用户未生成图表）"
            calc_info = f"\n包含前 {len(recent_calculations)} 次计算结果" if recent_calculations else "\n未包含计算结果"
            messagebox.showinfo("导出成功", f"数据已导出到: {filename}\n共导出 {len(operators)} 条记录{chart_info}{calc_info}")
            self._update_status(f"数据导出成功: {len(operators)} 条记录", "success")
            
            return True
            
        except Exception as e:
            logger.error(f"导出数据失败: {e}")
            messagebox.showerror("导出失败", f"导出数据时出现错误：\n{str(e)}")
            self._update_status("数据导出失败", "error")
            return False
    
    def _get_current_and_recent_calculations(self) -> List[Dict]:
        """获取当前用户计算结果和数据库中的历史计算记录"""
        combined_calculations = []
        
        try:
            # 1. 尝试从概览面板获取当前实际的计算结果
            if self.overview_panel and hasattr(self.overview_panel, 'main_window') and self.overview_panel.main_window:
                main_window = self.overview_panel.main_window
                if hasattr(main_window, 'panels') and 'analysis' in main_window.panels:
                    calc_panel = main_window.panels['analysis']
                    
                    # 获取当前的计算结果
                    current_results = self._extract_current_calculation_results(calc_panel)
                    if current_results:
                        combined_calculations.extend(current_results)
                        logger.info(f"从计算面板获取到 {len(current_results)} 条当前计算结果")
            
            # 2. 从数据库获取历史计算记录作为补充
            db_calculations = self.db_manager.get_calculation_history(limit=4)
            if db_calculations:
                # 避免重复：检查是否已经包含相同的计算结果
                for db_calc in db_calculations:
                    is_duplicate = False
                    for existing_calc in combined_calculations:
                        # 简单的重复检测：同一干员、相同参数
                        if (existing_calc.get('operator_name') == db_calc.get('operator_name') and
                            existing_calc.get('parameters') == db_calc.get('parameters')):
                            is_duplicate = True
                            break
                    
                    if not is_duplicate:
                        combined_calculations.append(db_calc)
                
                logger.info(f"从数据库补充了 {len(db_calculations)} 条历史计算记录")
            
            # 3. 限制总数量不超过10条，优先保留当前结果
            if len(combined_calculations) > 10:
                combined_calculations = combined_calculations[:10]
            
            logger.info(f"最终获取到 {len(combined_calculations)} 条计算记录用于导出")
            
        except Exception as e:
            logger.error(f"获取计算结果失败: {e}")
            # 后备方案：只使用数据库记录
            combined_calculations = self.db_manager.get_calculation_history(limit=4)
        
        return combined_calculations
    
    def _extract_current_calculation_results(self, calc_panel) -> List[Dict]:
        """从计算面板提取当前的计算结果"""
        current_results = []
        
        try:
            analysis_mode = getattr(calc_panel, 'analysis_mode', None)
            if not analysis_mode:
                return current_results
            
            mode = analysis_mode.get() if hasattr(analysis_mode, 'get') else str(analysis_mode)
            
            if mode == "single":
                # 单干员模式
                current_operator = getattr(calc_panel, 'current_operator', None)
                if current_operator:
                    # 获取当前计算参数
                    parameters = {
                        'enemy_def': getattr(calc_panel.enemy_def_var, 'get', lambda: 0)(),
                        'enemy_mdef': getattr(calc_panel.enemy_mdef_var, 'get', lambda: 0)(),
                        'calc_mode': getattr(calc_panel.calc_mode_var, 'get', lambda: 'basic_damage')(),
                    }
                    
                    # 获取当前计算结果
                    results = {
                        'dps': float(calc_panel.dps_result_var.get()) if hasattr(calc_panel, 'dps_result_var') else 0,
                        'dph': float(calc_panel.dph_result_var.get()) if hasattr(calc_panel, 'dph_result_var') else 0,
                        'armor_break': calc_panel.armor_break_var.get() if hasattr(calc_panel, 'armor_break_var') else '未知',
                    }
                    
                    # 如果有有效的计算结果，添加到列表
                    if results['dps'] > 0 or results['dph'] > 0:
                        calc_record = {
                            'operator_name': current_operator['name'],
                            'calculation_type': '单干员计算',
                            'parameters': parameters,
                            'results': results,
                            'created_at': datetime.now().isoformat()
                        }
                        current_results.append(calc_record)
                        logger.info(f"提取到单干员计算结果: {current_operator['name']}")
            
            elif mode == "multi":
                # 多干员模式
                multi_results = getattr(calc_panel, 'multi_comparison_results', None)
                if multi_results:
                    # 获取当前计算参数
                    parameters = {
                        'enemy_def': getattr(calc_panel.enemy_def_var, 'get', lambda: 0)(),
                        'enemy_mdef': getattr(calc_panel.enemy_mdef_var, 'get', lambda: 0)(),
                        'calc_mode': getattr(calc_panel.calc_mode_var, 'get', lambda: 'basic_damage')(),
                        'calc_mode_display': calc_panel._get_calc_mode_display_name(getattr(calc_panel.calc_mode_var, 'get', lambda: 'basic_damage')()) if hasattr(calc_panel, '_get_calc_mode_display_name') else '基础伤害计算'
                    }
                    
                    # 构建详细表格数据
                    detailed_table = []
                    max_dps = 0
                    total_dps = 0
                    max_efficiency = 0
                    
                    for operator_name, result in multi_results.items():
                        detailed_table.append({
                            '干员名称': operator_name,
                            '职业类型': result.get('class_type', ''),
                            '攻击类型': result.get('atk_type', ''),
                            '攻击力': result.get('atk', ''),
                            '攻击速度': result.get('atk_speed', ''),
                            '生命值': result.get('hp', ''),
                            '部署费用': result.get('cost', ''),
                            'DPS': result.get('dps', 0),
                            'DPH': result.get('dph', 0),
                            '破甲线': result.get('armor_break', ''),
                            '性价比': result.get('cost_efficiency', 0)
                        })
                        
                        # 统计数据
                        dps_val = float(result.get('dps', 0))
                        max_dps = max(max_dps, dps_val)
                        total_dps += dps_val
                        max_efficiency = max(max_efficiency, float(result.get('cost_efficiency', 0)))
                    
                    avg_dps = total_dps / len(multi_results) if multi_results else 0
                    
                    # 构建多干员对比结果
                    comparison_results = {
                        'detailed_table': detailed_table,
                        'max_dps': max_dps,
                        'avg_dps': avg_dps,
                        'max_efficiency': max_efficiency
                    }
                    
                    calc_record = {
                        'operator_name': f"多干员对比({len(multi_results)}个)",
                        'calculation_type': f'多干员对比计算 ({len(multi_results)} 个干员)',
                        'parameters': parameters,
                        'results': comparison_results,
                        'created_at': datetime.now().isoformat()
                    }
                    
                    current_results.append(calc_record)
                    logger.info(f"提取到多干员对比结果: {len(multi_results)} 个干员")
            
        except Exception as e:
            logger.error(f"提取当前计算结果失败: {e}")
        
        return current_results
    
    def _generate_export_charts(self, operators: List[Dict[str, Any]], base_filename: str) -> List[str]:
        """
        此方法已废弃，改为使用用户实际生成的图表
        """
        logger.warning("_generate_export_charts方法已废弃，现在使用用户实际生成的图表")
        return []
    
    def _export_csv_with_charts_and_calculations(self, operators: List[Dict[str, Any]], filename: str, chart_paths: List[str], recent_calculations: List[Dict]):
        """导出CSV文件并附带图表和计算结果"""
        # 导出CSV数据
        with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
            if operators:
                fieldnames = operators[0].keys()
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(operators)
        
        # 创建计算结果CSV文件
        if recent_calculations:
            calc_filename = filename.replace('.csv', '_计算结果.csv')
            with open(calc_filename, 'w', newline='', encoding='utf-8') as calc_csvfile:
                calc_fieldnames = ['计算序号', '计算类型', '计算时间', '计算参数', '对比汇总', '干员名称', '职业类型', '攻击类型', '攻击力', '攻击速度', '生命值', '部署费用', 'DPS', 'DPH', '破甲线', '性价比']
                calc_writer = csv.DictWriter(calc_csvfile, fieldnames=calc_fieldnames)
                calc_writer.writeheader()
                
                for i, calc in enumerate(recent_calculations, 1):
                    try:
                        # 解析参数和结果
                        parameters = calc.get('parameters', {})
                        results = calc.get('results', {})
                        
                        # 检查是否是多干员对比计算
                        if '多干员对比' in calc.get('calculation_type', '') and 'detailed_table' in results:
                            # 处理多干员对比的详细表格
                            detailed_table = results['detailed_table']
                            if detailed_table:
                                # 写入标题行
                                calc_writer.writerow({
                                    '计算序号': f"计算 {i}",
                                    '计算类型': calc.get('calculation_type', '未知'),
                                    '计算时间': str(calc.get('created_at', ''))[:19],
                                    '计算参数': f"敌防{parameters.get('enemy_def', 0)}, 敌法抗{parameters.get('enemy_mdef', 0)}, {parameters.get('calc_mode_display', '未知模式')}",
                                    '对比汇总': f"最大DPS {results.get('max_dps', 0):.2f}, 平均DPS {results.get('avg_dps', 0):.2f}, 最大性价比 {results.get('max_efficiency', 0):.2f}",
                                    '干员名称': '',
                                    '职业类型': '',
                                    '攻击类型': '',
                                    '攻击力': '',
                                    '攻击速度': '',
                                    '生命值': '',
                                    '部署费用': '',
                                    'DPS': '',
                                    'DPH': '',
                                    '破甲线': '',
                                    '性价比': ''
                                })
                                
                                # 写入每个干员的详细数据
                                for row_idx, row in enumerate(detailed_table):
                                    calc_writer.writerow({
                                        '计算序号': f"  干员 {row_idx + 1}",
                                        '计算类型': '',
                                        '计算时间': '',
                                        '计算参数': '',
                                        '对比汇总': '',
                                        '干员名称': row.get('干员名称', ''),
                                        '职业类型': row.get('职业类型', ''),
                                        '攻击类型': row.get('攻击类型', ''),
                                        '攻击力': row.get('攻击力', ''),
                                        '攻击速度': f"{float(row.get('攻击速度', 0)):.1f}",
                                        '生命值': row.get('生命值', ''),
                                        '部署费用': row.get('部署费用', ''),
                                        'DPS': f"{float(row.get('DPS', 0)):.2f}",
                                        'DPH': f"{float(row.get('DPH', 0)):.2f}",
                                        '破甲线': row.get('破甲线', ''),
                                        '性价比': f"{float(row.get('性价比', 0)):.2f}"
                                    })
                                
                                # 写入空行分隔
                                calc_writer.writerow({fieldname: '' for fieldname in calc_fieldnames})
                        else:
                            # 单干员计算记录
                            # 构建参数字符串
                            param_strs = []
                            if 'enemy_def' in parameters:
                                param_strs.append(f"敌防{parameters['enemy_def']}")
                            if 'enemy_mdef' in parameters:
                                param_strs.append(f"敌法抗{parameters['enemy_mdef']}")
                            if 'attack_type' in parameters:
                                param_strs.append(f"攻击类型{parameters['attack_type']}")
                            
                            calc_writer.writerow({
                                '计算序号': f"计算 {i}",
                                '计算类型': calc.get('calculation_type', '未知'),
                                '计算时间': str(calc.get('created_at', ''))[:19],
                                '计算参数': ' | '.join(param_strs),
                                '对比汇总': '',
                                '干员名称': calc.get('operator_name', '未知'),
                                '职业类型': '',
                                '攻击类型': '',
                                '攻击力': '',
                                '攻击速度': '',
                                '生命值': '',
                                '部署费用': '',
                                'DPS': f"{results.get('dps', 0):.2f}" if 'dps' in results else '-',
                                'DPH': f"{results.get('dph', 0):.2f}" if 'dph' in results else '-',
                                '破甲线': str(results.get('armor_break', '-')) if 'armor_break' in results else '-',
                                '性价比': ''
                            })
                        
                    except Exception as e:
                        logger.warning(f"导出计算记录 {i} 失败: {e}")
        
        # 创建说明文件
        readme_path = filename.replace('.csv', '_说明.txt')
        with open(readme_path, 'w', encoding='utf-8') as f:
            f.write("干员数据导出说明\n")
            f.write("=" * 30 + "\n\n")
            f.write(f"数据文件: {os.path.basename(filename)}\n")
            f.write(f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"干员数量: {len(operators)}\n")
            f.write(f"计算记录: {len(recent_calculations)} 条\n\n")
            
            if chart_paths:
                f.write("包含的图表文件:\n")
                for i, chart_path in enumerate(chart_paths, 1):
                    f.write(f"{i}. {os.path.basename(chart_path)}\n")
            
            if recent_calculations:
                f.write("\n计算结果文件:\n")
                calc_filename = filename.replace('.csv', '_计算结果.csv')
                f.write(f"计算结果详情: {os.path.basename(calc_filename)}\n")
    
    def _export_csv_with_charts(self, operators: List[Dict[str, Any]], filename: str, chart_paths: List[str]):
        """导出CSV文件并附带图表"""
        # 保持原有方法以确保向后兼容
        self._export_csv_with_charts_and_calculations(operators, filename, chart_paths, [])
    
    def _export_excel_with_charts_and_calculations(self, operators: List[Dict[str, Any]], filename: str, chart_paths: List[str], recent_calculations: List[Dict]):
        """导出Excel文件并插入图表和计算结果（美化版本）"""
        try:
            import pandas as pd
            from openpyxl import load_workbook
            from openpyxl.drawing.image import Image as OpenpyxlImage
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            # 创建Excel文件
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # 干员数据表
                df_operators = pd.DataFrame(operators)
                df_operators.to_excel(writer, sheet_name='干员数据', index=False)
                
                # 前4次计算结果表
                if recent_calculations:
                    calc_data = []
                    for i, calc in enumerate(recent_calculations, 1):
                        try:
                            parameters = calc.get('parameters', {})
                            results = calc.get('results', {})
                            
                            # 检查是否是多干员对比计算
                            if '多干员对比' in calc.get('calculation_type', '') and 'detailed_table' in results:
                                # 处理多干员对比的详细表格
                                detailed_table = results['detailed_table']
                                if detailed_table:
                                    # 添加标题行
                                    calc_data.append({
                                        '计算序号': f"🎯 计算 {i}",
                                        '计算类型': calc.get('calculation_type', '未知'),
                                        '计算时间': str(calc.get('created_at', ''))[:19],
                                        '计算参数': f"🛡️敌防{parameters.get('enemy_def', 0)} | 🔮敌法抗{parameters.get('enemy_mdef', 0)} | ⚙️{parameters.get('calc_mode_display', '未知模式')}",
                                        '对比汇总': f"📊最大DPS {results.get('max_dps', 0):.2f} | 📈平均DPS {results.get('avg_dps', 0):.2f} | 💰最大性价比 {results.get('max_efficiency', 0):.2f}",
                                        '干员名称': '',
                                        '职业类型': '',
                                        '攻击类型': '',
                                        '攻击力': '',
                                        '攻击速度': '',
                                        '生命值': '',
                                        '部署费用': '',
                                        'DPS': '',
                                        'DPH': '',
                                        '破甲线': '',
                                        '性价比': ''
                                    })
                                    
                                    # 添加每个干员的详细数据
                                    for row_idx, row in enumerate(detailed_table):
                                        calc_data.append({
                                            '计算序号': f"  📋 干员 {row_idx + 1}",
                                            '计算类型': '',
                                            '计算时间': '',
                                            '计算参数': '',
                                            '对比汇总': '',
                                            '干员名称': f"👤 {row.get('干员名称', '')}",
                                            '职业类型': f"🎯 {row.get('职业类型', '')}",
                                            '攻击类型': f"⚔️ {row.get('攻击类型', '')}",
                                            '攻击力': f"💪 {row.get('攻击力', '')}",
                                            '攻击速度': f"⚡ {float(row.get('攻击速度', 0)):.1f}",
                                            '生命值': f"❤️ {row.get('生命值', '')}",
                                            '部署费用': f"💰 {row.get('部署费用', '')}",
                                            'DPS': f"🔥 {float(row.get('DPS', 0)):.2f}",
                                            'DPH': f"💥 {float(row.get('DPH', 0)):.2f}",
                                            '破甲线': f"🛡️ {row.get('破甲线', '')}",
                                            '性价比': f"📊 {float(row.get('性价比', 0)):.2f}"
                                        })
                                    
                                    # 添加空行分隔
                                    calc_data.append({col: '' for col in ['计算序号', '计算类型', '计算时间', '计算参数', '对比汇总', '干员名称', '职业类型', '攻击类型', '攻击力', '攻击速度', '生命值', '部署费用', 'DPS', 'DPH', '破甲线', '性价比']})
                            else:
                                # 单干员计算记录
                                calc_record = {
                                    '计算序号': f"🎯 计算 {i}",
                                    '计算类型': f"📊 {calc.get('calculation_type', '未知')}",
                                    '计算时间': str(calc.get('created_at', ''))[:19],
                                    '计算参数': '',
                                    '对比汇总': '',
                                    '干员名称': f"👤 {calc.get('operator_name', '未知')}",
                                    '职业类型': '',
                                    '攻击类型': '',
                                    '攻击力': '',
                                    '攻击速度': '',
                                    '生命值': '',
                                    '部署费用': '',
                                    'DPS': f"🔥 {results.get('dps', 0):.2f}" if 'dps' in results else '➖',
                                    'DPH': f"💥 {results.get('dph', 0):.2f}" if 'dph' in results else '➖',
                                    '破甲线': f"🛡️ {str(results.get('armor_break', '➖'))}" if 'armor_break' in results else '➖',
                                    '性价比': ''
                                }
                                
                                # 构建参数字符串
                                param_strs = []
                                if 'enemy_def' in parameters:
                                    param_strs.append(f"🛡️敌防{parameters['enemy_def']}")
                                if 'enemy_mdef' in parameters:
                                    param_strs.append(f"🔮敌法抗{parameters['enemy_mdef']}")
                                if 'attack_type' in parameters:
                                    param_strs.append(f"⚔️{parameters['attack_type']}")
                                calc_record['计算参数'] = ' | '.join(param_strs)
                                
                                calc_data.append(calc_record)
                            
                        except Exception as e:
                            logger.warning(f"处理计算记录 {i} 失败: {e}")
                    
                    if calc_data:
                        df_calc = pd.DataFrame(calc_data)
                        df_calc.to_excel(writer, sheet_name='📈 计算结果详情', index=False)
                else:
                    # 如果没有计算记录，创建说明表
                    no_calc_info = [{
                        '📌 说明': '当前没有计算记录',
                        '💡 提示': '请在应用中进行计算分析后再导出',
                        '🔧 操作指南': '1. 在数据分析页面选择干员 → 2. 设置计算参数 → 3. 点击立即计算 → 4. 重新导出'
                    }]
                    df_no_calc = pd.DataFrame(no_calc_info)
                    df_no_calc.to_excel(writer, sheet_name='📈 计算结果详情', index=False)
                
                # 图表说明表
                if chart_paths:
                    chart_info = []
                    for i, chart_path in enumerate(chart_paths, 1):
                        chart_info.append({
                            '📊 序号': f"图表 {i}",
                            '📋 图表名称': os.path.basename(chart_path),
                            '📁 文件路径': chart_path,
                            '📝 说明': '🎨 用户在图表对比面板中生成的分析图表',
                            '⏰ 生成时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                            '📏 尺寸': '标准分析图表 (600x400)',
                            '🎯 用途': '数据可视化分析'
                        })
                    
                    df_charts = pd.DataFrame(chart_info)
                    df_charts.to_excel(writer, sheet_name='📊 图表说明', index=False)
                else:
                    # 如果没有图表，创建说明表
                    no_chart_info = [{
                        '📌 说明': '当前没有用户生成的图表',
                        '💡 提示': '请在图表对比面板中选择干员并生成图表后再导出',
                        '🔧 操作指南': '1. 切换到图表对比页面 → 2. 选择干员 → 3. 选择图表类型 → 4. 生成图表 → 5. 重新导出'
                    }]
                    df_no_charts = pd.DataFrame(no_chart_info)
                    df_no_charts.to_excel(writer, sheet_name='📊 图表说明', index=False)
            
            # 美化Excel格式
            try:
                wb = load_workbook(filename)
                
                # 定义样式
                header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
                header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                data_font = Font(name='微软雅黑', size=10)
                center_alignment = Alignment(horizontal='center', vertical='center')
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # 美化每个工作表
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    
                    # 设置标题行样式
                    for cell in ws[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = center_alignment
                        cell.border = border
                    
                    # 设置数据行样式
                    for row in ws.iter_rows(min_row=2):
                        for cell in row:
                            cell.font = data_font
                            cell.border = border
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    # 自动调整列宽
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        
                        adjusted_width = min(max_length + 2, 50)  # 最大宽度50
                        ws.column_dimensions[column_letter].width = adjusted_width
                    
                    # 冻结首行
                    ws.freeze_panes = 'A2'
                
                # 插入图表图片到Excel
                if chart_paths:
                    try:
                        # 创建图表展示工作表
                        if '🎨 图表展示' not in wb.sheetnames:
                            ws_charts = wb.create_sheet('🎨 图表展示')
                        else:
                            ws_charts = wb['🎨 图表展示']
                        
                        # 添加标题并美化
                        title_cell = ws_charts.cell(row=1, column=1)
                        title_cell.value = "📊 用户生成的图表展示"
                        title_cell.font = Font(name='微软雅黑', size=16, bold=True, color='2F5597')
                        
                        subtitle_cell = ws_charts.cell(row=2, column=1)
                        subtitle_cell.value = f"📈 共 {len(chart_paths)} 个图表 | ⏰ 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                        subtitle_cell.font = Font(name='微软雅黑', size=12, color='666666')
                        
                        # 插入图表
                        row = 4
                        for i, chart_path in enumerate(chart_paths, 1):
                            if os.path.exists(chart_path):
                                try:
                                    # 添加图表标题
                                    chart_name = os.path.basename(chart_path).replace('.png', '')
                                    chart_title_cell = ws_charts.cell(row=row, column=1)
                                    chart_title_cell.value = f"🎯 图表 {i}: {chart_name}"
                                    chart_title_cell.font = Font(name='微软雅黑', size=14, bold=True, color='2F5597')
                                    row += 1
                                    
                                    # 插入图片
                                    img = OpenpyxlImage(chart_path)
                                    img.width = 650  # 调整图片大小
                                    img.height = 450
                                    ws_charts.add_image(img, f'A{row}')
                                    row += 28  # 为下一个图表留出足够空间
                                    
                                    logger.info(f"成功插入美化图表到Excel: {chart_name}")
                                    
                                except Exception as e:
                                    logger.warning(f"插入图表 {chart_path} 失败: {e}")
                                    # 添加错误说明
                                    error_cell = ws_charts.cell(row=row, column=1)
                                    error_cell.value = f"❌ 图表 {i} 插入失败: {os.path.basename(chart_path)}"
                                    error_cell.font = Font(name='微软雅黑', size=12, color='FF0000')
                                    row += 2
                    
                        # 美化图表展示工作表
                        ws_charts.column_dimensions['A'].width = 80
                        
                    except Exception as e:
                        logger.warning(f"创建图表展示工作表失败: {e}")
                
                wb.save(filename)
                wb.close()
                
                logger.info(f"美化Excel文件已保存，包含 {len(chart_paths)} 个用户生成的图表和 {len(recent_calculations)} 条计算记录")
                
            except Exception as e:
                logger.warning(f"美化Excel格式失败: {e}")
                logger.info(f"基础Excel文件已保存，包含 {len(chart_paths)} 个用户生成的图表和 {len(recent_calculations)} 条计算记录")
                    
        except ImportError:
            # 如果没有pandas，使用基础方法
            logger.warning("pandas未安装，使用CSV格式导出")
            csv_filename = filename.replace('.xlsx', '.csv')
            self._export_csv_with_charts_and_calculations(operators, csv_filename, chart_paths, recent_calculations)
        except Exception as e:
            logger.error(f"导出Excel时出错: {e}")
            raise
    
    def _export_excel_with_charts(self, operators: List[Dict[str, Any]], filename: str, chart_paths: List[str]):
        """导出Excel文件并插入图表"""
        # 保持原有方法以确保向后兼容
        self._export_excel_with_charts_and_calculations(operators, filename, chart_paths, [])
    
    def _export_html_with_charts_and_calculations(self, operators: List[Dict[str, Any]], filename: str, chart_paths: List[str], recent_calculations: List[Dict]):
        """导出HTML报告并内嵌图表和计算结果"""
        import base64
        
        html_content = """
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>干员数据分析报告</title>
    <style>
        body {{ font-family: 'Microsoft YaHei', Arial, sans-serif; margin: 20px; background-color: #f5f5f5; }}
        .container {{ max-width: 1200px; margin: 0 auto; background: white; padding: 20px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
        h1, h2 {{ color: #333; border-bottom: 2px solid #4CAF50; padding-bottom: 10px; }}
        .summary {{ background: #e8f5e8; padding: 15px; border-radius: 5px; margin: 20px 0; }}
        table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
        th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
        th {{ background-color: #4CAF50; color: white; }}
        tr:nth-child(even) {{ background-color: #f2f2f2; }}
        .chart-section {{ margin: 30px 0; text-align: center; }}
        .chart-image {{ max-width: 100%; height: auto; border: 1px solid #ddd; border-radius: 5px; margin: 10px 0; }}
        .calc-section {{ margin: 30px 0; }}
        .calc-item {{ background: #f8f9fa; border-left: 4px solid #007bff; padding: 15px; margin: 10px 0; }}
        .footer {{ text-align: center; margin-top: 30px; color: #666; font-size: 12px; }}
    </style>
</head>
<body>
    <div class="container">
        <h1>🎯 干员数据分析报告</h1>
        
        <div class="summary">
            <h2>📊 数据概览</h2>
            <p><strong>报告生成时间:</strong> {timestamp}</p>
            <p><strong>干员总数:</strong> {total_operators} 个</p>
            <p><strong>计算记录:</strong> {calculation_count} 条</p>
        </div>
        
        <h2>📋 详细数据表</h2>
        <table>
            <thead>
                <tr>
                    {table_headers}
                </tr>
            </thead>
            <tbody>
                {table_rows}
            </tbody>
        </table>
        
        {calculations_section}
        
        {charts_section}
        
        <div class="footer">
            <p>本报告由明日方舟伤害分析器自动生成</p>
        </div>
    </div>
</body>
</html>
        """
        
        # 计算统计数据
        total_operators = len(operators)
        
        # 生成表格
        if operators:
            headers = list(operators[0].keys())
            table_headers = ''.join([f"<th>{header}</th>" for header in headers])
            
            table_rows = ""
            for op in operators:
                row = "<tr>"
                for header in headers:
                    value = op.get(header, '')
                    row += f"<td>{value}</td>"
                row += "</tr>"
                table_rows += row
        else:
            table_headers = "<th>无数据</th>"
            table_rows = "<tr><td>无数据</td></tr>"
        
        # 生成计算结果部分
        calculations_section = ""
        if recent_calculations:
            calculations_section = f"<h2>📈 前4次计算结果详情 (共{len(recent_calculations)}条)</h2>"
            calculations_section += "<div class='calc-section'>"
            
            for i, calc in enumerate(recent_calculations, 1):
                try:
                    operator_name = calc.get('operator_name', '未知干员')
                    calc_type = calc.get('calculation_type', '未知计算')
                    created_at = str(calc.get('created_at', ''))[:19]
                    
                    parameters = calc.get('parameters', {})
                    results = calc.get('results', {})
                    
                    # 构建参数字符串
                    param_details = []
                    if 'enemy_def' in parameters:
                        param_details.append(f"敌人防御: {parameters['enemy_def']}")
                    if 'enemy_mdef' in parameters:
                        param_details.append(f"敌人法抗: {parameters['enemy_mdef']}")
                    if 'attack_type' in parameters:
                        param_details.append(f"攻击类型: {parameters['attack_type']}")
                    
                    # 构建结果字符串
                    result_details = []
                    if 'dps' in results:
                        result_details.append(f"DPS: {results['dps']:.2f}")
                    if 'dph' in results:
                        result_details.append(f"单发伤害: {results['dph']:.2f}")
                    if 'total_damage' in results:
                        result_details.append(f"总伤害: {results['total_damage']:.0f}")
                    if 'hps' in results:
                        result_details.append(f"HPS: {results['hps']:.2f}")
                    
                    calculations_section += f"""
                    <div class="calc-item">
                        <h4>计算 {i}: {operator_name} - {calc_type}</h4>
                        <p><strong>计算时间:</strong> {created_at}</p>
                        <p><strong>计算参数:</strong> {' | '.join(param_details) if param_details else '无参数信息'}</p>
                        <p><strong>计算结果:</strong> {' | '.join(result_details) if result_details else '无结果信息'}</p>
                    </div>
                    """
                except Exception as e:
                    logger.warning(f"处理计算记录 {i} 失败: {e}")
                    calculations_section += f"""
                    <div class="calc-item">
                        <h4>计算 {i}: 数据解析失败</h4>
                        <p>无法解析此条计算记录</p>
                    </div>
                    """
            
            calculations_section += "</div>"
        else:
            calculations_section = """
            <h2>📈 计算结果</h2>
            <div class='calc-section'>
                <div class="calc-item">
                    <p>当前没有计算记录。请在应用中进行一些计算分析。</p>
                </div>
            </div>
            """
        
        # 生成图表部分
        charts_section = ""
        if chart_paths:
            charts_section = f"<h2>📊 用户生成的图表 (共{len(chart_paths)}个)</h2>"
            for chart_path in chart_paths:
                if os.path.exists(chart_path):
                    try:
                        # 将图片转换为base64编码内嵌到HTML中
                        with open(chart_path, 'rb') as img_file:
                            img_data = base64.b64encode(img_file.read()).decode()
                            chart_name = os.path.basename(chart_path).replace('.png', '')
                            charts_section += f"""
                            <div class="chart-section">
                                <h3>{chart_name}</h3>
                                <img src="data:image/png;base64,{img_data}" class="chart-image" alt="{chart_name}">
                            </div>
                            """
                    except Exception as e:
                        logger.warning(f"处理图表 {chart_path} 失败: {e}")
        else:
            charts_section = """
            <h2>📊 图表展示</h2>
            <div class="chart-section">
                <p>当前没有用户生成的图表。请在图表对比面板中生成图表。</p>
            </div>
            """
        
        # 填充模板
        html_final = html_content.format(
            timestamp=datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            total_operators=total_operators,
            calculation_count=len(recent_calculations),
            table_headers=table_headers,
            table_rows=table_rows,
            calculations_section=calculations_section,
            charts_section=charts_section
        )
        
        # 保存HTML文件
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(html_final)
    
    def _export_html_with_charts(self, operators: List[Dict[str, Any]], filename: str, chart_paths: List[str]):
        """导出HTML报告并内嵌图表"""
        # 保持原有方法以确保向后兼容
        self._export_html_with_charts_and_calculations(operators, filename, chart_paths, [])
    
    def export_all_data_to_excel(self, operators: List[Dict[str, Any]], filename: str = None) -> bool:
        """
        导出所有数据到Excel
        
        Args:
            operators: 干员数据列表
            filename: 导出文件路径
            
        Returns:
            是否导出成功
        """
        try:
            if not operators:
                messagebox.showwarning("警告", "没有数据可导出")
                return False
            
            if filename is None:
                filename = filedialog.asksaveasfilename(
                    title="导出Excel文件",
                    defaultextension=".xlsx",
                    filetypes=[
                        ("Excel 文件", "*.xlsx"),
                        ("所有文件", "*.*")
                    ]
                )
                
            if not filename:
                return False
            
            try:
                import pandas as pd
                from datetime import datetime
                
                # 显示进度
                self._update_status("正在生成Excel文件...", "info")
                
                # 创建Excel写入器
                with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                    # 1. 干员数据表
                    df_operators = pd.DataFrame(operators)
                    df_operators.to_excel(writer, sheet_name='干员数据', index=False)
                    
                    # 2. 计算记录表
                    try:
                        calc_records = self.db_manager.get_calculation_history(limit=1000)
                        if calc_records:
                            # 处理计算记录数据
                            processed_calc = []
                            for record in calc_records:
                                processed_record = {
                                    'ID': record.get('id'),
                                    '干员名称': record.get('operator_name', '未知'),
                                    '计算类型': record.get('calculation_type', ''),
                                    '创建时间': record.get('created_at', ''),
                                    '参数': str(record.get('parameters', {})),
                                    '结果': str(record.get('results', {}))
                                }
                                processed_calc.append(processed_record)
                            
                            df_calc = pd.DataFrame(processed_calc)
                            df_calc.to_excel(writer, sheet_name='计算记录', index=False)
                    except Exception as e:
                        logger.warning(f"导出计算记录失败: {e}")
                    
                    # 3. 导入记录表
                    try:
                        import_records = self.db_manager.get_import_records(limit=1000)
                        if import_records:
                            df_import = pd.DataFrame(import_records)
                            df_import.to_excel(writer, sheet_name='导入记录', index=False)
                    except Exception as e:
                        logger.warning(f"导出导入记录失败: {e}")
                    
                    # 4. 统计摘要表
                    try:
                        stats = self.db_manager.get_statistics_summary()
                        stats_data = [
                            {'统计项目': '干员总数', '数值': stats.get('total_operators', 0)},
                            {'统计项目': '导入记录总数', '数值': stats.get('total_imports', 0)},
                            {'统计项目': '计算记录总数', '数值': stats.get('total_calculations', 0)},
                            {'统计项目': '今日计算次数', '数值': stats.get('today_calculations', 0)},
                        ]
                        
                        # 添加职业分布
                        class_dist = stats.get('class_distribution', {})
                        for class_type, count in class_dist.items():
                            stats_data.append({'统计项目': f'{class_type}职业干员', '数值': count})
                        
                        df_stats = pd.DataFrame(stats_data)
                        df_stats.to_excel(writer, sheet_name='统计摘要', index=False)
                    except Exception as e:
                        logger.warning(f"导出统计摘要失败: {e}")
                
                # 记录导出操作
                self.db_manager.record_import(
                    import_type='excel_export',
                    file_name=os.path.basename(filename),
                    record_count=len(operators),
                    status='success'
                )
                
                messagebox.showinfo("导出成功", f"Excel文件已导出到: {filename}\n包含多个数据表，共 {len(operators)} 个干员数据")
                self._update_status("Excel导出成功", "success")
                
                return True
                
            except ImportError:
                messagebox.showerror("错误", "需要安装pandas和openpyxl库才能导出Excel文件\n请运行: pip install pandas openpyxl")
                return False
                
        except Exception as e:
            logger.error(f"导出Excel失败: {e}")
            messagebox.showerror("导出失败", f"导出Excel时出现错误：\n{str(e)}")
            self._update_status("Excel导出失败", "error")
            return False
    
    def quick_import_dialog(self, parent_window) -> Dict[str, Any]:
        """
        快速导入对话框
        
        Args:
            parent_window: 父窗口
            
        Returns:
            导入结果字典
        """
        try:
            filename = filedialog.askopenfilename(
                title="快速导入数据文件",
                filetypes=[
                    ("所有支持格式", "*.xlsx;*.xls;*.json;*.csv"),
                    ("Excel 文件", "*.xlsx;*.xls"),
                    ("JSON 文件", "*.json"),
                    ("CSV 文件", "*.csv"),
                    ("所有文件", "*.*")
                ]
            )
            
            if not filename:
                return {'success': False, 'cancelled': True}
            
            file_ext = filename.lower().split('.')[-1]
            
            if file_ext in ['xlsx', 'xls']:
                return self.import_excel_data(filename)
            elif file_ext == 'json':
                return self.import_json_data(filename)
            elif file_ext == 'csv':
                return self.import_csv_data(filename)
            else:
                messagebox.showwarning("警告", "不支持的文件格式")
                return {'success': False, 'error': '不支持的文件格式'}
            
        except Exception as e:
            logger.error(f"快速导入失败: {e}")
            error_msg = f"快速导入失败：\n{str(e)}"
            messagebox.showerror("导入失败", error_msg)
            return {'success': False, 'error': str(e)}
    
    def export_excel_with_current_charts(self, operators: List[Dict[str, Any]], current_charts: List[Dict] = None, filename: str = None) -> bool:
        """
        导出Excel文件并包含用户生成的图表
        
        Args:
            operators: 干员数据列表
            current_charts: 用户生成的图表列表
            filename: 导出文件路径
            
        Returns:
            是否导出成功
        """
        try:
            if not operators:
                messagebox.showwarning("警告", "没有数据可导出")
                return False
            
            if filename is None:
                filename = filedialog.asksaveasfilename(
                    title="导出Excel文件(包含用户生成图表)",
                    defaultextension=".xlsx",
                    filetypes=[
                        ("Excel 文件", "*.xlsx"),
                        ("所有文件", "*.*")
                    ]
                )
                
            if not filename:
                return False
            
            base_filename = filename.rsplit('.', 1)[0]
            
            # 保存用户生成的图表为图片文件
            chart_paths = []
            if current_charts:
                chart_paths = self._save_current_charts_as_images(current_charts, base_filename)
                logger.info(f"保存了 {len(chart_paths)} 个用户生成的图表")
            else:
                logger.info("没有用户生成的图表可导出")
            
            # 导出Excel文件
            self._export_excel_with_charts(operators, filename, chart_paths)
            
            # 记录导出操作
            self.db_manager.record_import(
                import_type='excel_export_with_user_charts',
                file_name=os.path.basename(filename),
                record_count=len(operators),
                status='success'
            )
            
            # 显示成功消息
            chart_info = f"\n包含 {len(chart_paths)} 个用户生成的图表" if chart_paths else "\n未包含图表（用户未生成图表）"
            messagebox.showinfo("导出成功", f"Excel文件已导出到: {filename}\n共导出 {len(operators)} 条记录{chart_info}")
            self._update_status(f"Excel导出成功: {len(operators)} 条记录", "success")
            
            return True
            
        except Exception as e:
            logger.error(f"导出Excel失败: {e}")
            messagebox.showerror("导出失败", f"导出Excel时出现错误：\n{str(e)}")
            self._update_status("Excel导出失败", "error")
            return False
    
    def _save_current_charts_as_images(self, current_charts: List[Dict], base_filename: str) -> List[str]:
        """将当前图表保存为图片文件"""
        chart_paths = []
        
        try:
            for i, chart_info in enumerate(current_charts):
                try:
                    figure = chart_info.get('figure')
                    title = chart_info.get('title', f'图表_{i+1}')
                    chart_type = chart_info.get('type', 'unknown')
                    
                    if figure:
                        # 清理文件名中的非法字符
                        safe_title = "".join(c for c in title if c.isalnum() or c in (' ', '-', '_')).rstrip()
                        chart_path = f"{base_filename}_{safe_title}.png"
                        
                        # 保存图表
                        figure.savefig(chart_path, dpi=300, bbox_inches='tight', facecolor='white')
                        chart_paths.append(chart_path)
                        
                        logger.info(f"已保存图表: {chart_path}")
                        
                except Exception as e:
                    logger.warning(f"保存图表 {i} 失败: {e}")
                    
        except Exception as e:
            logger.error(f"保存当前图表失败: {e}")
        
        return chart_paths 
    
    def export_excel_with_current_charts_and_calculations(self, operators: List[Dict[str, Any]], current_charts: List[Dict] = None, current_calculations: List[Dict] = None, filename: str = None) -> bool:
        """
        导出Excel文件并包含用户生成的图表和当前计算结果
        
        Args:
            operators: 干员数据列表
            current_charts: 用户生成的图表列表
            current_calculations: 用户当前的计算结果列表
            filename: 导出文件路径
            
        Returns:
            是否导出成功
        """
        try:
            if not operators:
                messagebox.showwarning("警告", "没有数据可导出")
                return False
            
            if filename is None:
                filename = filedialog.asksaveasfilename(
                    title="导出Excel文件(包含用户生成图表和计算结果)",
                    defaultextension=".xlsx",
                    filetypes=[
                        ("Excel 文件", "*.xlsx"),
                        ("所有文件", "*.*")
                    ]
                )
                
            if not filename:
                return False
            
            base_filename = filename.rsplit('.', 1)[0]
            
            # 保存用户生成的图表为图片文件
            chart_paths = []
            if current_charts:
                chart_paths = self._save_current_charts_as_images(current_charts, base_filename)
                logger.info(f"保存了 {len(chart_paths)} 个用户生成的图表")
            else:
                logger.info("没有用户生成的图表可导出")
            
            # 使用传入的当前计算结果，而不是从数据库重新获取
            if current_calculations is None:
                # 如果没有传入计算结果，则尝试获取
                current_calculations = self._get_current_and_recent_calculations()
            
            # 导出Excel文件，包含图表和当前计算结果
            self._export_excel_with_charts_and_calculations(operators, filename, chart_paths, current_calculations)
            
            # 记录导出操作
            self.db_manager.record_import(
                import_type='excel_export_with_user_data',
                file_name=os.path.basename(filename),
                record_count=len(operators),
                status='success'
            )
            
            # 显示成功消息
            chart_info = f"\n包含 {len(chart_paths)} 个用户生成的图表" if chart_paths else "\n未包含图表（用户未生成图表）"
            calc_info = f"\n包含 {len(current_calculations)} 条当前计算结果" if current_calculations else "\n未包含计算结果"
            messagebox.showinfo("导出成功", f"Excel文件已导出到: {filename}\n共导出 {len(operators)} 条记录{chart_info}{calc_info}")
            self._update_status(f"Excel导出成功: {len(operators)} 条记录，{len(current_calculations)} 条计算结果", "success")
            
            return True
            
        except Exception as e:
            logger.error(f"导出Excel失败: {e}")
            messagebox.showerror("导出失败", f"导出Excel时出现错误：\n{str(e)}")
            self._update_status("Excel导出失败", "error")
            return False